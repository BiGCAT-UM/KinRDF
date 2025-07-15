##This script has been tested with Python version 3.13.5

##Load library to report which Python version was used for reproducability
import sys
print("Python version: ", sys.version)
print("Version info: ", sys.version_info)

##Load library to check and set local path easier
import os
from os.path import join

##Define variables for folder names:
checkedData = "/KineticsData" ##For curated data collection
testData = "/TestData" ##For all test data
validationData = "/ValidateData" ##To test one file at a time from test data

##Select folder name to apply script on:
subfolderLocation = checkedData ##Adapt this statement if needed
##TODO: include different setup in 3 separate GitHub Actions.

#Update the path to where the data is stored.
dir_code = os.getcwd()
if dir_code.endswith('KinRDF'):
  dataFolder = dir_code + subfolderLocation
else:
  parentPath = os.path.dirname(dir_code) #go up one directory
  dir_code = parentPath
  dataFolder = dir_code + subfolderLocation
os.chdir(dataFolder) ##Update directory to folder with data
dir_code = os.getcwd()

#Load library for regex to test for structure of content
import re

##Lists to store data in
ListTotal = []
ListSER_ID = []
ListSER_ID_type = []
ListUniprot = [] 
ListLinkUniprot = [] 
ListProteinsLinks = [] 
ListEnsembl = []
ListRheaID = []
ListRheaID_type = []
ListLinkRheaID = [] 
ListInteractionLinks = [] 
ListSubstrate = []
ListSubstrateIDs = []
ListSubstrateLinks = [] 
ListKm = []
ListKcat = []
ListKcatKm = []
List_pH = []
ListTemperature = []
ListAdditionalConditions = []
ListOrganism = []
ListPMID = []
ListDatabase = []
ListProv = []
ListQC = []
ListErrors = []
ListCuration = [] 

##Header String to compare input data against template (in exact order)
headers = 'EnzymePW	ApprovedEnzymeName	E.C.number	Uniprot	Ensembl	RheaID	CHEBIID	SusbtrateName	Km (mM)	Kcat (s-1)	Kcat/Km (M-1s-1)	pH	Temperature (Celcius)	AdditionalConditions	Organism	PMID	Database\n'
headersList = ['EnzymePW', 'ApprovedEnzymeName', 'E.C.number', 'Uniprot', 'Ensembl', 'RheaID', 'CHEBIID', 'SusbtrateName', 'Km (mM)', 'Kcat (s-1)', 'Kcat/Km (M-1s-1)', 'pH', 'Temperature (Celcius)', 'AdditionalConditions', 'Organism', 'PMID', 'Database']

##Import library to read xlsx files:
import openpyxl
##Import library to read .csv files:
import csv

##Read in files with kinetics data.
#count = 0
countSER = 1
for (dirname, dirs, files) in os.walk('.'):
	for filename in files:
	  if filename.endswith('.txt') :
	    thefile = os.path.join(dirname,filename)
	    try:
	      f = open(filename, "r", encoding="utf8").readlines() ##Read files with UTF8 encoding specified
	      header_line = f.pop(0)
	    except UnicodeDecodeError:
	      ListQC.append("File contains not UTF8 encoded characters, CHECK: "+ filename + '\n')
	    if (header_line == headers):
	      for line in f:
	        SER_Name = "SER:" + str(countSER)
	        ListTotal.append(SER_Name + '\t' + line.strip('\n') + '\t' + filename) ##Add filename to list to easily find curation point.
	        countSER += 1
	    else:
	        ListQC.append("File contains wrong column names, CHECK: "+ filename + '\n')
	  elif filename.endswith('.xlsx') :
	    thefile = os.path.join(dirname,filename)
	    try:
	      wb_obj = openpyxl.load_workbook(filename) ##Files are read with UTF8 encoding as standard
	      sheet_obj = wb_obj.active
	      header_line = [cell.value for cell in sheet_obj[1]]
	      rows_iter = sheet_obj.iter_rows(min_col = 1, min_row = 2, max_col = sheet_obj.max_column, max_row = sheet_obj.max_row)
	      values = [[cell.value for cell in row] for row in rows_iter]
	      entries = range(len(values))
	    except UnicodeDecodeError:
	      ListQC.append("File contains not UTF8 encoded characters, CHECK: "+ filename + '\n')
	    if (header_line == headersList):
	      for n in entries:
	        SER_Name = "SER:" + str(countSER)
	        line = '\t'.join(str(element) for element in values[n])
	        ListTotal.append(SER_Name + '\t' + line  + '\t' + filename) ##Add filename to list to easily find curation point.
	        countSER += 1
	    else:
	        ListQC.append("File contains wrong column names, CHECK: "+ filename + '\n')
	  elif filename.endswith('.csv') :
	    List_rawcsvdata = []
	    thefile = os.path.join(dirname,filename)
	    try:
	      f = open(filename, newline='', encoding='utf-8') ##Open the csv file
	      file_data = csv.reader(f, delimiter=',', quotechar='"') ##Read the data with delimiter ','  and keep cells content withint double quotations.
	      for line in file_data:
	        List_rawcsvdata.append('\t'.join(line)) ##Convert data to .tsv structure for further processing.
	      header_line = List_rawcsvdata[0] + '\n'
	      if (header_line == headers):
	        for line in List_rawcsvdata[1:len(List_rawcsvdata)]: ##Skipping the first line, which is the header (column names) of the file.
	          SER_Name = "SER:" + str(countSER)
	          ListTotal.append(SER_Name + '\t' + line.strip('\n') + '\t' + filename) ##Add filename to list to easily find curation point.
	          countSER += 1
	    except UnicodeDecodeError:
	      ListQC.append("File contains not UTF8 encoded characters, CHECK: "+ filename + '\n')
	  else:
	    ListQC.append("File extension could not be read: "+ filename + '\n')

###Cleaning up data:

##Remove spaces between prefixes and IDs for RHEA (column 6) and ChEBI (column 7)
ListTotal = [w.replace('RHEA: ', '') for w in ListTotal]
ListTotal = [w.replace('Rhea: ', '') for w in ListTotal]
ListTotal = [w.replace('rhea: ', '') for w in ListTotal]
ListTotal = [w.replace('CHEBI: ', '') for w in ListTotal]
ListTotal = [w.replace('Chebi: ', '') for w in ListTotal]
ListTotal = [w.replace('ChEBI: ', '') for w in ListTotal]
ListTotal = [w.replace('chebi: ', '') for w in ListTotal]

##Remove text within parenthesis using regex:
##TODO: find regex to remove double brackets, e.g. '((3R,5S)-1-pyrroline-3-hydroxy-5-carboxylate)'
ListTotal = [re.sub(r"[\(\[].*?[\)\]]","",x) for x in ListTotal] ##Everything in brackets

##Replace common prefixes for harmonized data structure:
ListTotal = [w.replace('RHEA:', '') and w.replace('Rhea:', '') and w.replace('rhea:', '') for w in ListTotal]
ListTotal = [w.replace('CHEBI:', '') and w.replace('Chebi:', '') and w.replace('ChEBI:', '') and w.replace('chebi:', '') for w in ListTotal]

##Replace items with backslash as empty value
ListTotal = [re.sub("\\\\","",x) for x in ListTotal]
##Replace items with forward as empty value
ListTotal = [re.sub("/","",x) for x in ListTotal]

##Replace 'EC:' for enzyme nomenclature IDs if available
ListTotal = [w.replace('EC: ', '')  for w in ListTotal]
ListTotal = [w.replace('EC:', '') for w in ListTotal]
ListTotal = [w.replace('EC', '') for w in ListTotal]

##Replace misspelled organism name for humans (Homo sapien)
ListTotal = [w.replace('Homo sapien\t', 'Homo sapiens\t') for w in ListTotal]
##Replace not common abbreviations for species
ListTotal = [w.replace('house mouse', 'mouse') for w in ListTotal] ##Mouse
ListTotal = [w.replace('House Mouse', 'mouse') for w in ListTotal] ##Mouse
ListTotal = [w.replace('brown rat', 'rat') for w in ListTotal] ##Rat
ListTotal = [w.replace('Brown Rat', 'rat') for w in ListTotal] ##Rat
ListTotal = [w.replace('wild boar or pig', 'pig') for w in ListTotal] ##Pig

##Remove lines without ChEBI, UniProt, or Rhea (Excel files include 'None' for empty values)
for item in ListTotal[:]:
	a = item.split('\t')
	#Check if one of the necessary values is missing!!
	if (a[0].strip()=='-')|(a[4].strip()=='-')|(a[6].strip()=='-')|(a[7].strip()=='-') | (a[0].strip()=='NA')|(a[4].strip()=='NA')|(a[6].strip()=='NA')|(a[7].strip()=='NA') | (a[0].strip()=='None')|(a[4].strip()=='None')|(a[6].strip()=='None')|(a[7].strip()=='None') | (a[0].strip()=='')|(a[4].strip()=='')|(a[6].strip()=='')|(a[7].strip()==''):
	  ListCuration.append("CURATION: missing data for: "+ a[0] + ' substrate: ' + a[7] + ' enzyme: ' + a[4] + ' reaction: ' + a[6] + ' ' + a[18] + '\n')
	  ListTotal.remove(item)

##Remove lines without any provenance (either PMID or Database are needed)
### '-'
for item in ListTotal[:]:
	a = item.split('\t')
	if ((a[16].strip()=='-')&(a[17].strip()=='-'))|((a[16].strip()=='-')&(a[17].strip()=='NA'))|((a[16].strip()=='-')&(a[17].strip()=='')|((a[16].strip()=='-')&(a[17].strip()=='None'))): #Check if both values are missing!!
	  ListCuration.append("CURATION: Provenance for data is missing, check original data for: "+ a[0] + " publication: " + a[16] + ' database: ' + a[17] + ' ' + a[18] + '\n')  
	  ListTotal.remove(item)

### 'NA'
for item in ListTotal[:]:
	a = item.split('\t')	  
	if ((a[16].strip()=='NA')&(a[17].strip()=='NA'))|((a[16].strip()=='NA')&(a[17].strip()=='-'))|((a[16].strip()=='NA')&(a[17].strip()==''))|((a[16].strip()=='NA')&(a[17].strip()=='None')):  #Check if both values are NA!!
	  ListCuration.append("CURATION: Provenance for data is missing, check original data for: "+ a[0] + " publication: " + a[16] + ' database: ' + a[17] + ' ' + a[18] + '\n')  
	  ListTotal.remove(item)

### ''
for item in ListTotal[:]:
	a = item.split('\t')
	if ((a[16].strip()=='')&(a[17].strip()==''))|((a[16].strip()=='')&(a[17].strip()=='-'))|((a[16].strip()=='')&(a[17].strip()=='NA'))|((a[16].strip()=='')&(a[17].strip()=='None')):  #Check if both values are empty!!
	  ListCuration.append("CURATION: Provenance for data is missing, check original data for: "+ a[0] + " publication: " + a[16] + ' database: ' + a[17] + ' ' + a[18] + '\n')  
	  ListTotal.remove(item)

### 'None' (from .xlsx files)
for item in ListTotal[:]:
	a = item.split('\t')
	if ((a[16].strip()=='None')&(a[17].strip()=='None'))|((a[16].strip()=='None')&(a[17].strip()=='-'))|((a[16].strip()=='None')&(a[17].strip()=='NA'))|((a[16].strip()=='None')&(a[17].strip()=='')):  #Check if both values are None!!
	  ListCuration.append("CURATION: Provenance for data is missing, check original data for: "+ a[0] + " publication: " + a[16] + ' database: ' + a[17] + ' ' + a[18] + '\n')  
	  ListTotal.remove(item)

##Print total number of lines found in files:
ListQC.append("Total lines read: "+ str(len(ListTotal)) + '\n')

##Regex:
pattern_chebi = '^(CHEBI:)?\\d+$'
pattern_uniprot = '^([A-N,R-Z][0-9][A-Z][A-Z, 0-9][A-Z, 0-9][0-9])|([O,P,Q][0-9][A-Z, 0-9][A-Z, 0-9][A-Z, 0-9][0-9])(\\.\\d+)?|([A-N,R-Z][0-9][A-Z][A-Z, 0-9][A-Z, 0-9][0-9][A-Z][A-Z, 0-9][A-Z, 0-9][0-9])$'
pattern_rhea = '^(RHEA:)?\\d{5}$'

##SER_number
countSER = 0
for itemSERX in ListTotal[:]:
	a = itemSERX.split('\t')
	pattern = '[a-zA-Z]+:[0-9]'
	result = re.match(pattern, a[0])
	result_chebi =  re.match(pattern_chebi, a[7].strip())
	result_uniprot = re.match(pattern_uniprot, a[4].strip())
	result_rhea = re.match(pattern_rhea, a[6].strip())
	if (result_chebi is not None) & (result_rhea is not None) & (result_uniprot is not None): ##check against REGEX
	  ListSER_ID.append(a[0].strip( ) + '\t' + 'dc:identifier' + ' ' + 'SER:'+ a[7].strip( ) + '-' + a[4].strip( ) + '-' + a[6].strip( )) ##Trim RHEA/CHEBI in fromt of IDs if available.
	  ListSER_ID.append(a[0].strip() + '\t' + 'sio:SIO_000028 ' + ' ' +  a[0].strip( ) + '_substrate' + ', '  + a[0].strip( )+ '_enzyme' + ', ' + a[0].strip( )+ '_reaction') #Add the 'has part' relationship, so we can link the IDs to that later.
	  ListSER_ID.append(a[0].strip() + '\t' + 'sio:SIO_000008 ' + ' ' +  a[0].strip( ) + '_measurement' ) #Add the 'has attribute' relationship, so we can link the values to that later.
	  countSER += 1
	elif(result_chebi is not None) & (result_rhea is None)  & (result_uniprot is not None): ##check against REGEX; if RHEA is not available, use reaction formula
	  ListSER_ID.append(a[0].strip( ) + '\t' + 'dc:identifier' + ' ' + 'SER:'+ a[7].strip( ) + '-' + a[4].strip( ) + '-XXXXX')
	  ListSER_ID.append(a[0].strip() + '\t' + 'sio:SIO_000028 ' + ' ' +  a[0].strip( ) + '_substrate' + ', '  + a[0].strip( )+ '_enzyme' + ', ' + a[0].strip( )+ '_reaction')
	  ListSER_ID.append(a[0].strip() + '\t' + 'sio:SIO_000008 ' + ' ' +  a[0].strip( ) + '_measurement' ) #Add the 'has attribute' relationship, so we can link the values to that later.
	  countSER += 1
	elif((result_rhea is not None) & (result_uniprot is not None) & ((all(x.isnumeric() or x.isspace() for x in a[7].strip()))|(',' in a[7].strip())|(';' in a[7].strip()))): ##Substrate column contains multiple values
	    ListErrors.append("CHECK: Multiple values for ChEBI IDs detected, check original data for: "+ a[0] + " : " + a[7] + ' ' + a[18] + '\n')
	else:
	  ListCuration.append("CURATION: missing data for: "+ a[0] + ' substrate: ' + a[7] + ' enzyme: ' + a[4] + ' reaction: ' + a[6] + ' ' + a[18] + '\n')
	  ListTotal.remove(itemSERX)

##Print total number of lines found in files, after removing data without SEP-ID:
ListQC.append("Lines remaining without missing SER info: "+ str(len(ListTotal)) + '\n')

##Print total number of SEP-IDs, for which measurements are available:
ListQC.append("Data format for SER correctly loaded for " + str(countSER) + " Substrate, Enzyme, and Reaction IDs. \n\n")  

#### Template Version 2.x
#### [0]=ERPX_number [1]=EnzymePW, [2]=ApprovedEnzymeName, [3]=EC_Number, [4]=Uniprot , 
#### [5]=Ensembl, [6]=RheaID, [7]=CHEBIID, [9] = SubstrateName, [9]=Km, [10]=Kcat, [11]=Kcat/Km, [12]=pH, [13]=Temperature, 
#### [14]=AdditionalConditions, [15]=Organism, [16]=PMID, [17]=Database				

##EnzymePW 
##No regex or count defined, since the names of Proteins can be very diverse!
for itemEnzymePW in ListTotal:
	b = itemEnzymePW.split('\t')
	if (b[1].strip()=='-')|(b[1].strip()=='NA')|(b[1].strip()=='')|(b[1].strip()=='None')|(b[4].strip()=='-')|(b[4].strip()=='NA')|(b[4].strip()=='')|(b[4].strip()=='None'): #Check if one of the necessary values is missing!!
	  continue
	else:
	  ListUniprot.append('uniprot:' + b[4].strip( ) + '\t' + 'rdfs:label' + ' "' + b[1].strip( ) + '"^^xsd:string')

###Approved Ennzyme names (added in spreadsheet for curation, not needed in RDF model)
# ##[2]=ApprovedEnzymeName
# for itemApprovedEnzymeName in ListTotal:
	# c = itemApprovedEnzymeName.split('\t')
	# ListApprovedEnzymeName.append(b[0] + ' ' + 'rdfs:label' + ' ' + c[2])
	# for items in ListApprovedEnzymeName: 
		# if '-' in items:
			# ListApprovedEnzymeName.remove(items)

##EC_Numbers
countECs = 0
for itemEC_Number in ListTotal:
	d = itemEC_Number.split('\t')
	pattern = '^\\d+\\.-\\.-\\.-|\\d+\\.\\d+\\.-\\.-|\\d+\\.\\d+\\.\\d+\\.-|\\d+\\.\\d+\\.\\d+\\.(n)?\\d+$'
	result = re.match(pattern, d[3].strip())
	if (d[3].strip()=='-')|(d[3].strip()=='NA')|(d[3].strip()=='')|(d[4].strip()=='-')|(d[4].strip()=='NA')|(d[4].strip()=='')|(d[4].strip()=='None'): #Check if necessary value is missing!!
	  continue
	elif result:
	  ListUniprot.append('uniprot:' + d[4].strip( )  + '\t' + 'wp:bdbEnzymeNomenclature' + ' ECcode:' + d[3].strip( ))
	  countECs = countECs + 1
	else:
	  ListErrors.append("CHECK: Data format for wp:bdbEnzymeNomenclature unknown, check original data for: "+ d[0] + ' : ' + d[3] + ' ' + d[18] + '\n')
    
ListQC.append("Data format for 'wp:bdbEnzymeNomenclature correctly loaded for " + str(countECs) + " EC IDs. \n\n")  

##Uniprot IDs
countProteins = 0
##Connect measurement data to UniProt IDs:
countProteinsLinks = 0
for itemUniprot in ListTotal:
	e = itemUniprot.split('\t')
	if(e[4].strip()=='-')|(e[4].strip()=='NA')|(e[4].strip()=='')|(e[4].strip()=='None'): #Check if necessary value is missing!!
	  continue
	else:
	  ListUniprot.append('uniprot:' + e[4].strip( ) + '\t'  + 'rdf:type' + ' ' + "wp:Protein")
	  ##WP IRIs:
	  ListUniprot.append('uniprot:' + e[4].strip( )  + '\t' + 'wp:bdbUniprot' + ' uniprot:' + e[4].strip( )) 
	  ##Uniprot IRIs for UniProt RDF link: "uniprotkb:P05067 a up:Protein ;"
	  ListLinkUniprot.append('uniprot:' + e[4].strip( ) + '\t' + 'bioregistry:hasDbXref' + ' ' + 'uniprotkb:'  + e[4].strip() + '.' )
	  countProteins = countProteins + 1
	  ListProteinsLinks.append(e[0].strip( ) + '_enzyme' + '\t' + 'rdfs:subClassOf' + ' ' + 'uniprot:' + e[4].strip( ) + '.')
	  countProteinsLinks = countProteinsLinks + 1

##Print total number of UniProt IDs:    
ListQC.append("Data format for wp:bdbUniprot correctly loaded for " + str(countProteins) + " UniProt Protein IDs. \n\n")  
##Print total number of UniProt IDs linked to measurements:    
ListQC.append("Data format for Uniprot correctly linked for " + str(countProteinsLinks) + " measurement IDs. \n\n")  

##Ensembl IDs
countGenes = 0
for itemEnsembl in ListTotal:
  f = itemEnsembl.split('\t')
  pattern = '^ENS[A-Z]*[FPTG]\\d{11}$' #Pattern is: ^ENS[A-Z]*[FPTG]\d{11}$' ; need to escape backslash! 
  result = re.match(pattern, f[5].strip( ) )
  if (f[5].strip()=='-')|(f[5].strip()=='NA')|(f[5].strip()=='')|(f[5].strip()=='None')|(f[4].strip()=='-')|(f[4].strip()=='NA')|(f[4].strip()=='')|(f[4].strip()=='None'): #Check if one of the necessary values is missing!!
    continue
  elif result: ##check against REGEX
    ListUniprot.append('uniprot:' + f[4].strip( ) + '\t' + 'wp:bdbEnsembl' + ' En_id:' + f[5].strip( ))
    countGenes = countGenes + 1
  else:
    ListErrors.append("CHECK: Data format for wp:bdbEnsembl unknown, check original data for: "+ f[0] + ' : ' + f[5] + ' ' + f[18] + '\n')

##Print total number of Ensembl IDs:    
ListQC.append("Data format for wp:bdbEnsembl correctly loaded for " + str(countGenes) + " gene IDs. \n\n")

##Define type, extension of WP vocabulary:		
#for itemSERX in ListTotal:
#	a = itemSERX.split('\t')
#	ListSER_ID_type.append(a[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData')

##RHEA IDs
countRhea = 0
countEquation = 0
##Connect measurement data to Rhea IDs:
countRheaLinks = 0
for itemRheaID in ListTotal:
	g = itemRheaID.split('\t')
	pattern_rhea = '^(RHEA:)?\\d{5}$'
	result_rhea = re.match(pattern_rhea, g[6].strip())
	if (g[6].strip()=='-')|(g[6].strip()=='NA')|(g[6].strip()=='')|(g[6].strip()=='None'): #Check if one of the necessary values is missing!!
	  ListRheaID_type.append(g[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData') ##To make sure statement ends with type
	  ListErrors.append("CHECK: Data format for Rhea unknown, check original data for: "+ g[0] + ' : ' + g[6] + ' ' + g[18] + '\n')
	  continue
	elif ('R-HSA-' in g[6].strip()):
	  ##TODO: add conversion to Rhea IDs in the future from BridgeDb mapping file!
	  ListRheaID_type.append(g[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData') ##To make sure statement ends with type
	  ListErrors.append("CHECK: Data format for Rhea contains Reactome IDs, check original data for: "+ g[0] + ' : ' + g[6] + ' ' + g[18] + '\n')
	  continue
	elif(result_rhea): ##regex check
	  if ( "rhea" in g[6].strip().lower() ) : ##If Rhea is part of the ID structure.
	    ListRheaID.append( g[6].strip( ) + '\t' + 'wp:bdbRhea' + ' ' + g[6].strip( ) ) 
	    ListLinkRheaID.append( g[6].strip( ) + '\t' + 'rdf:type' + ' ' + "wp:Interaction" + ' ;')	  
	    ListLinkRheaID.append( g[6].strip( ) + '\t' + 'rh:accession' + ' ' + g[6].strip( ) + '.' )	
	    ListRheaID_type.append(g[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData') ##To make sure statement ends with type
	    ListInteractionLinks.append(g[0].strip( ) + '_reaction' + '\t' + 'rdfs:subClassOf' + ' ' + g[6].strip( ) + '.')
	    countRheaLinks = countRheaLinks + 1
	  else:  ##If Rhea is NOT part of the ID structure.
	    ListRheaID.append('RHEA:' + g[6].strip( ) + '\t' + 'wp:bdbRhea'  + ' RHEA:' + g[6].strip( ) ) 
	    ListLinkRheaID.append('RHEA:' + g[6].strip( ) + '\t' + 'rdf:type' + ' ' + "wp:Interaction" + ' ;')	  
	    ListLinkRheaID.append('RHEA:' + g[6].strip( ) + '\t' + 'rh:accession' + ' RHEA:' + g[6].strip( ) + '.' )	
	    ListRheaID_type.append(g[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData') ##To make sure statement ends with type
	    ListInteractionLinks.append(g[0].strip( ) + '_reaction' + '\t' + 'rdfs:subClassOf' + ' ' + 'RHEA:' + g[6].strip( ) + '.')
	    countRheaLinks = countRheaLinks + 1
	  countRhea = countRhea +1
	elif ('+' in g[6])|('=' in g[6]): #if no Rhea is available, but there is a reaction equation.
	  ListRheaID.append(g[0].strip( ) + '\t' + 'rh:equation' + ' "' + g[6].strip( ) + '"^^xsd:string') ##Missing directional info!!
	  ListRheaID_type.append(g[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData') ##To make sure statement ends with type
	  countEquation = countEquation + 1   
	elif (result_rhea is None): ##Rhea doesn't match regex, but isn't empty  
	  ListRheaID_type.append(g[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData') ##To make sure statement ends with type
	  ListErrors.append("CHECK: Data format for Rhea unknown, check original data for: "+ g[0] + ' : ' + g[6] + ' ' + g[18] + '\n')
	else: #if no Rhea is available
	  ListRheaID_type.append(g[0].strip( ) + '\t' + 'rdf:type ' + 'wp:InteractionData') ##To make sure statement ends with type
	  ListErrors.append("CHECK: Data format for Rhea unknown, check original data for: "+ g[0] + ' : ' + g[6] + ' ' + g[18] +  '\n')

##Print total number of Rhea IDs:    
ListQC.append("Data format for wp:bdbRhea correctly loaded for " + str(countRhea) + " rhea interaction IDs. \n\n")  
ListQC.append("Data format for rh:equation correctly loaded for " + str(countEquation) + " reaction formulas. \n\n")  
##Print total number of Linked Rhea IDs to measurements:    
ListQC.append("Data format for Rhea correctly linked for " + str(countRheaLinks) + " measurement IDs. \n\n")  

##CHEBI IDs
countSubstrates = 0
##Connect measurement data to ChEBI IDs:
countSubstratesLinks = 0
#ListSubstrateLinks
for itemSubstrate in ListTotal:
	h = itemSubstrate.split('\t')
	pattern_chebi = '^(CHEBI:)?\\d+$'
	result_chebi =  re.match(pattern_chebi, h[7].strip())
	if (h[7].strip()=='-')|(h[7].strip()=='NA')|(h[7].strip()=='')|(h[7].strip()=='None'): #Check if one of the necessary values is missing!!
	  continue
	elif(result_chebi):
	  if ( "chebi" in h[7].strip().lower() ) : ##If ChEBI is part of the ID structure.
	    ListSubstrateIDs.append( h[7].strip( ) + '\t'  + 'a' + ' ' + 'CHEBI:' + "23367" + ' ;') #molecular entity
	    ListSubstrateIDs.append( h[7].strip( ) + '\t'  + 'rdf:type' + ' ' + "wp:Metabolite" + ' ;')
	    ListSubstrateIDs.append( h[7].strip( ) + "\t" + "wp:bdbChEBI" + ' ' + h[7].strip( )+ '.')
	    countSubstrates = countSubstrates + 1
	    ListInteractionLinks.append(h[0].strip( ) + '_substrate' + '\t' + 'rdfs:subClassOf' + ' ' + h[7].strip( ) + '.')
	    countSubstratesLinks = countSubstratesLinks + 1
	  else: ##If ChEBI is NOT part of the ID structure.
	    ListSubstrateIDs.append('CHEBI:' + h[7].strip( ) + '\t'  + 'a' + ' ' + 'CHEBI:' + "23367" + ' ;') #molecular entity
	    ListSubstrateIDs.append('CHEBI:' + h[7].strip( ) + '\t'  + 'rdf:type' + ' ' + "wp:Metabolite" + ' ;')
	    ListSubstrateIDs.append('CHEBI:' + h[7].strip( ) + "\t" + "wp:bdbChEBI" + ' ' + 'CHEBI:' + h[7].strip( )+ '.')
	    countSubstrates = countSubstrates + 1
	    ListInteractionLinks.append(h[0].strip( ) + '_substrate' + '\t' + 'rdfs:subClassOf' + ' ' + 'CHEBI:' + h[7].strip( ) + '.')
	    countSubstratesLinks = countSubstratesLinks + 1
	else:
	  ListErrors.append("CHECK: Data format for CHEBI ID unknown, check original data for: "+ h[0] + " : " + h[7] + ' ' + h[18] + '\n')

##Print total number of ChEBI IDs:  
ListQC.append("Data format for wp:bdbChEBI correctly loaded for " + str(countSubstrates) + " substrate IDs. \n\n")  
##Print total number of Linked ChEBI IDs to measurements:    
ListQC.append("Data format for ChEBI correctly linked for " + str(countSubstratesLinks) + " measurement IDs. \n\n")  

###Substrate names (added in spreadsheet for curation, not needed in RDF model)
##To be updated if needed
# ##[8]=ApprovedEnzymeName
# for itemApprovedEnzymeName in ListTotal:
	# c = itemApprovedEnzymeName.split('\t')
	# ListApprovedEnzymeName.append(b[0] + ' ' + 'rdfs:label' + ' ' + c[2])
	# for items in ListApprovedEnzymeName: 
		# if '-' in items:
			# ListApprovedEnzymeName.remove(items)


##Regex:
pattern_float = r'^\d+\.\d+$' ##to test for numbers with dot-decimal separator

##Km
countKm = 0
for itemKm in ListTotal:
  i = itemKm.split('\t')
  i[9] = i[9].replace(',', '.') ##Replace comma decimal values with a dot decimal for consitency
  result_float =  re.match(pattern_float, i[9].strip())
  if(i[9].strip()=='-')|(i[9].strip()=='NA')|(i[9].strip()=='')|(i[9].strip()=='None') | (((i[17].strip()=='-')|(i[17].strip()=='NA')|(i[17].strip()=='')|(i[17].strip()=='None'))&((i[16].strip()=='-')|(i[16].strip()=='NA')|(i[16].strip()=='')|(i[16].strip()=='None'))): #Check if one of the necessary value && provenance are missing!!
    continue
  elif(result_float): ##Check if entry contains a number with decimal(s)
    ListKm.append(i[0].strip() + '_measurement' + '\t' + 'SER:hasKm' + ' "' + i[9].strip() + '"^^xsd:float')
    countKm = countKm + 1
  elif(i[9].strip().isnumeric()): ##Check if entry contains a number without decimal(s)
    ListKm.append(i[0].strip() + '_measurement' + '\t' + 'SER:hasKm' + ' "' + i[9].strip() + '"^^xsd:float')
    countKm = countKm + 1
  elif("e" in i[9].strip().lower()): ##Check if entry contains a scientific annotations 
    try:
      Scinumber = float(i[9].strip().lower())
      ListKm.append(i[0].strip() + '_measurement' + '\t' + 'SER:hasKm' + ' "' + str('%.08f' % Scinumber) + '"^^xsd:float')
      countKm = countKm + 1  
    except: ##To avoid issues with other strings containing an 'e'
      ListErrors.append("CHECK: Data format for Km data not a scientific number, check original data for: "+ i[0] + " : " + i[9] + ' ' + i[18] + '\n')
  else:
    ListErrors.append("CHECK: Data format for Km data not numeric, check original data for: "+ i[0] + " : " + i[9] + ' ' + i[18] + '\n')
	  
ListQC.append("Data format for Km correctly loaded for " + str(countKm) + " values. \n\n")  

##Kcat
countKcat = 0
for itemKcat in ListTotal:
  i = itemKcat.split('\t')
  i[10] = i[10].replace(',', '.') ##Replace comma decimal values with a dot decimal for consitency
  result_float =  re.match(pattern_float, i[10].strip())
  if(i[10].strip()=='-')|(i[10].strip()=='NA')|(i[10].strip()=='')|(i[10].strip()=='None') | (((i[17].strip()=='-')|(i[17].strip()=='NA')|(i[17].strip()=='')|(i[17].strip()=='None'))&((i[16].strip()=='-')|(i[16].strip()=='NA')|(i[16].strip()=='')|(i[16].strip()=='None'))): #Check if one of the necessary values is missing!!
    continue
  elif(result_float):
    ListKcat.append(i[0].strip() + '_measurement' + '\t' + 'SER:hasKcat' + ' "' + i[10].strip() + '"^^xsd:float') 
    countKcat = countKcat + 1
  elif(i[10].strip().isnumeric()):
    ListKcat.append(i[0].strip() + '_measurement' + '\t' + 'SER:hasKcat' + ' "' + i[10].strip() + '"^^xsd:float') 
    countKcat = countKcat + 1
  else:
    ListErrors.append("CHECK: Data format for Kcat data not numeric, check original data for: "+ i[0] + " : " + i[10] + ' ' + i[18] + '\n')

ListQC.append("Data format for Kcat correctly loaded for " + str(countKcat) + " values. \n\n") 

#KcatKm
countKcatKm = 0
for itemKcatKm in ListTotal:
  i = itemKcatKm.split('\t')
  i[11] = i[11].replace(',', '.') ##Replace comma decimal values with a dot decimal for consitency
  result_float =  re.match(pattern_float, i[11].strip())
  if(i[11].strip()=='-')|(i[11].strip()=='NA')|(i[11].strip()=='')|(i[11].strip()=='None') | (((i[17].strip()=='-')|(i[17].strip()=='NA')|(i[17].strip()=='')|(i[17].strip()=='None'))&((i[16].strip()=='-')|(i[16].strip()=='NA')|(i[16].strip()=='')|(i[16].strip()=='None'))): #Check if one of the necessary values is missing!!
    continue
  elif(result_float):
    ListKcatKm.append(i[0].strip() + '_measurement' + '\t' + 'SER:hasKcatKm' + ' "' + i[11].strip() + '"^^xsd:float') 
    countKcatKm = countKcatKm + 1	
  elif(i[11].strip().isnumeric()):
    ListKcatKm.append(i[0].strip() + '_measurement' + '\t' + 'SER:hasKcatKm' + ' "' + i[11].strip() + '"^^xsd:float') 
    countKcatKm = countKcatKm + 1	
  else:
    ListErrors.append("CHECK: Data format for KcatKm data not numeric, check original data for: "+ i[0] + " : " + i[11] + ' ' + i[18] + '\n')

ListQC.append("Data format for KcatKm correctly loaded for " + str(countKcatKm) + " values. \n\n") 

#pH##--> Add to measurement
count_pH = 0
for item_pH in ListTotal:
  i = item_pH.split('\t')
  i[12] = i[12].replace(',', '.') ##Replace comma decimal values with a dot decimal for consitency
  result_float =  re.match(pattern_float, i[12].strip())
  if(i[12].strip()=='-')|(i[12].strip()=='NA')|(i[12].strip()=='')|(i[12].strip()=='None') | (((i[17].strip()=='-')|(i[17].strip()=='NA')|(i[17].strip()=='')|(i[17].strip()=='None'))&((i[16].strip()=='-')|(i[16].strip()=='NA')|(i[16].strip()=='')|(i[16].strip()=='None'))) : #Check if one of the necessary values is missing!!
    continue
  elif(result_float):
    List_pH.append(i[0].strip( ) + '_measurement' + '\t' + 'SER:hasPh'+ ' "' + i[12].strip( ) + '"^^xsd:float')
    count_pH = count_pH + 1
  elif(i[12].strip().isnumeric()):
    List_pH.append(i[0].strip( ) + '_measurement' + '\t' + 'SER:hasPh'+ ' "' + i[12].strip( ) + '"^^xsd:float')
    count_pH = count_pH + 1
  else:
    ListErrors.append("CHECK: Data format for pH data not numeric, check original data for: "+ i[0] + " : " + i[12]  + ' ' + i[18] + '\n')

ListQC.append("Data format for pH correctly loaded for " + str(count_pH) + " values. \n\n") 
			
#Temperature##--> Add to measurement
countTemp = 0
for itemTemperature in ListTotal:
  i = itemTemperature.split('\t')
  i[13] = i[13].replace(',', '.') ##Replace comma decimal values with a dot decimal for consitency
  result_float =  re.match(pattern_float, i[13].strip())
  if(i[13].strip()=='-')|(i[13].strip()=='NA')|(i[13].strip()=='')|(i[13].strip()=='None') | (((i[17].strip()=='-')|(i[17].strip()=='NA')|(i[17].strip()=='')|(i[17].strip()=='None'))&((i[16].strip()=='-')|(i[16].strip()=='NA')|(i[16].strip()=='')|(i[16].strip()=='None'))): #Check if one of the necessary values is missing!!
    continue
  elif(result_float):
    ListTemperature.append(i[0].strip( ) + '_measurement' + '\t' + 'wdt:P2076' + ' "' + i[13].strip( ) + '"^^xsd:float') #Line from after 2020-01-17
    countTemp = countTemp + 1
  elif(i[13].strip().isnumeric()):
    ListTemperature.append(i[0].strip( ) + '_measurement' + '\t' + 'wdt:P2076' + ' "' + i[13].strip( ) + '"^^xsd:float') #Line from after 2020-01-17
    countTemp = countTemp + 1
  else:
    ListErrors.append("CHECK: Data format for Temp data not numeric, check original data for: "+ i[0] + " : " + i[13] + ' ' + i[18] + '\n')

ListQC.append("Data format for Temperature correctly loaded for " + str(countTemp) + " values. \n\n") 

#AdditionalConditions##--> Add to measurement
countConditions = 0
for itemAdditionalConditions in ListTotal:
	i = itemAdditionalConditions.split('\t')
	if(i[14].strip()=='-')|(i[14].strip()=='NA')|(i[14].strip()=='')|(i[14].strip()=='None') | (((i[17].strip()=='-')|(i[17].strip()=='NA')|(i[17].strip()=='')|(i[17].strip()=='None'))&((i[16].strip()=='-')|(i[16].strip()=='NA')|(i[16].strip()=='')|(i[16].strip()=='None'))): #Check if one of the necessary values is missing!!
	  continue
	else:
	  ListAdditionalConditions.append(i[0].strip( )  + '_measurement' + '\t' + 'rdfs:comment' + ' "' + i[14].strip('"') + '"@en')
	  countConditions = countConditions + 1
	  
ListQC.append("Data format for Additional Conditions correctly loaded for " + str(countConditions) + " values. \n\n") 

## Lists to check listed organism (Latin and common English names) for animals//vertebrates commonly used in WikiPathways:
ListCommonLatinOrganismsWP = ['Bos taurus', 'Canis familiaris', 'Danio rerio', 'Equus caballus', 'Gallus gallus', 'Homo sapiens', 'Mus musculus', 'Pan troglodytes', 'Rattus norvegicus', 'Sus scrofa', 'Escherichia coli']
ListCommonEnglishOrganismsWP = ['cow', 'dog', 'zebrafish', 'horse', 'chicken', 'human', 'mouse', 'chimpanzee', 'rat', 'pig', 'E. coli'] 
##TODO: add mouse, rat as separate entries (list with multiple values to compare against 1 key)

ListTaxonIDs = ['9913', '9615', '7955', '9796', '9031', '9606', '10090', '9598', '10116', '9823', '562']
#  ?pathway wp:organism ?organism .
#  filter(contains(str(?organism),"9606")) 

# using dictionary comprehension to convert lists to dictionary for conversion of English to Latin later.
Dict_CommonOrganismsWP = {ListCommonEnglishOrganismsWP[i]: ListCommonLatinOrganismsWP[i] for i in range(len(ListCommonEnglishOrganismsWP))}
Dict_TaxonIDs = {ListCommonLatinOrganismsWP[i]: ListTaxonIDs[i] for i in range(len(ListCommonLatinOrganismsWP))}

##Other data currently present in datasets (not modeled in RDF):
## Bacteria
# Alcaligenes sp. 
# Chlorobium limicola
# Rhoopseudomonas palustris
# Synechocystis sp.
# Pasteurella multocida
## Plants:
# Arabidopsis thaliana
# Pisum sativum
## Fungi:
# Saccharomyces cerevisiae
## Vertebrates
# Oryctolagus cuniculus (european rabbit)

#[15]=Organism			##--> Add to measurement
countOrganisms = 0
for itemOrganism in ListTotal:
	j = itemOrganism.split('\t')
	if(j[15].strip()=='-')|(j[15].strip()=='NA')|(j[15].strip()=='')|(j[15].strip()=='None') | (((i[17].strip()=='-')|(i[17].strip()=='NA')|(i[17].strip()=='')|(i[17].strip()=='None'))&((i[16].strip()=='-')|(i[16].strip()=='NA')|(i[16].strip()=='')|(i[16].strip()=='None'))): #Check if one of the necessary values is missing!!
	  continue
	elif((all(x.isalpha() or x.isspace() for x in j[15])|('.' in j[15]))): ##Check if Organism names only contains letters and spaces only; or a dot (for abbreviated names)
	  if(j[15].lower().strip().replace(" ", "") in [x.lower().replace(" ", "") for x in ListCommonLatinOrganismsWP]): ##check for latin name first
	    for key, value in Dict_CommonOrganismsWP.items(): ## Add names from structured data, not input data.
	      if value.strip().lower().replace(" ", "") == j[15].lower().strip().replace(" ", ""):
	        ListOrganism.append(j[0].strip( ) + '_measurement'  + '\t' + 'wp:organismName' + ' "' + Dict_CommonOrganismsWP[key] + '"^^xsd:string')
	        countOrganisms = countOrganisms + 1
	        for keys, values in Dict_TaxonIDs.items():
	          if keys.strip().lower().replace(" ", "") == j[15].lower().strip().replace(" ", ""):
	            ListOrganism.append(j[0].strip( ) + '_measurement'  + '\t' + 'wd:P703' + ' "' + Dict_TaxonIDs[keys] + '"') #found in taxon
	  elif(j[15].lower().strip().replace(" ", "") in [x.lower().replace(" ", "") for x in ListCommonEnglishOrganismsWP]):  ##Convert English to Latin name.
	    for key,value in Dict_CommonOrganismsWP.items():
	      if key.strip().lower().replace(" ", "") == j[15].lower().strip().replace(" ", ""):
	        ListOrganism.append(j[0].strip( ) + '_measurement'  + '\t' + 'wp:organismName' + ' "' + Dict_CommonOrganismsWP[key] + '"^^xsd:string')
	        for keys, values in Dict_TaxonIDs.items():
	          if keys.strip().lower().replace(" ", "") == value.strip().lower().replace(" ", ""):
	            ListOrganism.append(j[0].strip( ) + '_measurement'  + '\t' + 'wd:P703' + ' "' + Dict_TaxonIDs[keys] + '"') #found in taxon
	    countOrganisms = countOrganisms + 1
	  else:
	    ListErrors.append("CHECK: Name for Organism is not recognized, check original data for: "+ j[0] + " : " + j[15] + ' ' + j[18] + '\n')
	else:
	  ListErrors.append("CHECK: Data format for Organism name is not recognized, check original data for: "+ j[0] + " : " + j[15] + ' ' + j[18] + '\n')

ListQC.append("Data format for Organisms correctly loaded for " + str(countOrganisms) + " values. \n\n") 

##Provenance can come from two sources; a PMID, a database name, or both.
##At least one is needed to support the data  in the RDF.

#[16]=PMID	##--> Add to measurement	
#[17]=Database ##--> Add to measurement				
ListSupportedDatabases = ['brenda', 'sabio', 'guide to pharmacology', 'strenda', 'uniprot']

ListSupportedDatabasesAlternatives = ['brenda', 'sabio-rk', 'guide to pharmacology', 'strenda', 'uniprot'] ##If no alternative name is known (yet), use same name as official.

# using dictionary comprehension to convert lists to dictionary for conversion of English to Latin later.
Dict_SupportedDatabases = {ListSupportedDatabasesAlternatives[i]: ListSupportedDatabases[i] for i in range(len(ListSupportedDatabasesAlternatives))}

countRefs = 0
countProv = 0

for itemProv in ListTotal:
	p = itemProv.split('\t')
	if((p[16].strip()=='-')|(p[16].strip()=='NA')|(p[16].strip()=='')|(p[16].strip()=='None'))&((p[17].strip()=='-')|(p[17].strip()=='NA')|(p[17].strip()=='')|(p[17].strip()=='None')): ## if both are missing, do not include data.
	  continue
	####First scenario, both values are available and valid:
	##Option 1: Pubmed contains 1 value; database name contains 1 value
	elif ((p[16].isnumeric())&((all(x.isalpha() or x.isspace() for x in p[17])|((('-' in p[17])|((':' in p[17])))&((',' not in p[17])&(';' not in p[17])))))):  
	  ##Check if pubmed ID is numeric and if database names only contains letters (or spaces, or one bar for sabio-rk, or ':' for internal brenda IDs).:
	  splitProv = p[17].split(':') ##Doesn't trow an error if no ':' is present!
	  if(((p[17].strip().lower() not in ListSupportedDatabases)&(p[17].strip().lower() not in ListSupportedDatabasesAlternatives))&((splitProv[0].strip().lower() not in ListSupportedDatabases)&(splitProv[0].strip().lower() not in ListSupportedDatabasesAlternatives))): 
	    ##Check if database name is not valid, so only pubmed ID will be added!
	    ListProv.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " pubmed:" + p[16].strip( ))
	    countRefs += 1
	    if(p[17].strip() == '-'):
	      ##check if database is equally to '-', and ignore this value in QC report.
	      continue
	    else:
	      ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')  
	  elif((p[17].strip().lower() in ListSupportedDatabases)|(splitProv[0].strip().lower() in ListSupportedDatabases)): ##check for official name first
	    ListPMID.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " pubmed:" + p[16].strip( ))
	    countRefs += 1
	    for key, value in Dict_SupportedDatabases.items(): ## Add databases from structured data, not input data.
	      if ((value.strip().lower().replace(" ", "") == p[17].strip().lower().replace(" ", ""))|(value.strip().lower().replace(" ", "") == splitProv[0].strip().lower().replace(" ", ""))):
	        ListDatabase.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' "' + Dict_SupportedDatabases[key] + '"^^xsd:string')
	    countProv = countProv + 1
	  elif((p[17].strip().lower() in ListSupportedDatabasesAlternatives)|(splitProv[0].strip().lower() in ListSupportedDatabasesAlternatives)):  ##Convert alternative name to official name.
	    ListPMID.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " pubmed:" + p[16].strip( ))
	    countRefs += 1
	    for key in Dict_SupportedDatabases:
	      if ((key.strip().lower().replace(" ", "") == p[17].strip().lower().replace(" ", ""))|(key.strip().lower().replace(" ", "")  == splitProv[0].strip().lower().replace(" ", ""))): 
	        ListDatabase.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' "' + Dict_SupportedDatabases[key] + '"^^xsd:string')
	    countProv = countProv + 1  
	  else: ##For any other case Database name is not recognised
	    ListProv.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " pubmed:" + p[16].strip( ))
	    countRefs += 1
	    if(p[17].strip() == '-'):
	      continue
	    else:
	      ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')  
	##Option 2: Pubmed contains 1 value; database contains more than 1
	elif (p[16].isnumeric())&((';' in p[17])|(',' in p[17])|(':' in p[17])): ##Check if pubmed ID is numeric and database name contains multiple values semicolon or comma separated. Note: items with mulitple separators, e.g. with internal Brenda IDs will not be split.
	  ListPMID.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " pubmed:" + p[16].strip( ))
	  countRefs += 1
	  if(';' in p[17])&(',' not in p[17])&(':' not in p[17]):
	    p2 = p[17].split(';') ##Split multiple references in one line, semicolon split.
	  elif(',' in p[17])&(';' not in p[17])&(':' not in p[17]):
	    p2 = p[17].split(',') ##Split multiple references in one line, comma split.
	  else:
	    ListErrors.append("CHECK: Multiple separators for Database Provenance detected, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')  
	  try:
	    ListProvenance = []
	    p2 = [x.strip(' ').lower() for x in p2] ##strip whitespaces (if existing)
	    for z in [0,(len(p2)-1)]:
	      if(p2[z] in ListSupportedDatabases): ##check for original name first
	        p2[z] = '"' + p2[z] + '"^^xsd:string' ##add suffix for each item in list
	        ListProvenance.append(p2[z])
	      elif(p2[z] in ListSupportedDatabasesAlternatives):  ##Check to convert alternative name to official name.
	        for key in Dict_SupportedDatabases:
	          if key.lower() == p2[z].strip().lower():
	            p2[z] = Dict_SupportedDatabases[key] ##Convert alternative name to official to item
	            p2[z] = '"' + p2[z] + '"^^xsd:string'##add suffix for item
	            ListProvenance.append(p2[z])
	      else:
	        ListProvenance = []
	        ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')  
	  except NameError:
	    continue
	  ListDatabase.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' ' + ', '.join(ListProvenance))
	  try:
	    countProv = countProv + len(p2)
	  except NameError:
	    countProv = countProv
	##Option 3: Pubmed contains more than 1 value, database name contains 1 value
	elif ((';' in p[16])|(',' in p[16]))&(all(x.isalpha() or x.isspace() for x in p[17])):
	  if(';' in p[16]):
	    p3 = p[16].split(';') ##Split multiple references in one line.
	  elif(',' in p[16]):
	    p3 = p[16].split(',') ##Split multiple references in one line.
	  else:
	    ListErrors.append("CHECK: Multiple separators for Pubmed IDs detected, check original data for: "+ p[0] + " : " + p[16] + ' ' + p[18] + '\n')  
	  p3 = [x.strip(' ') for x in p3] ##strip whitespaces (if existing)
	  for y in [0,(len(p3)-1)]:
	    if(p3[y].isnumeric()):
	      p3 = ['pubmed:' + s for s in p3] ##add prefix for each item in list
	      ListPMID.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " " + ', '.join(p3))
	  countRefs = countRefs + len(p3)
	  if(p[17].strip().lower() in ListSupportedDatabases): ##check for latin name first
	    ListDatabase.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' "' + p[17].strip( ).lower() + '"^^xsd:string')
	    countProv = countProv + 1
	  else:
	    ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')
	##Option 4: Pubmed contains more than 1 value, database name contains more than 1 value    
	elif ((';' in p[16])|(',' in p[16]))&((';' in p[17])|(',' in p[16])):
	  if(';' in p[16])&(';' in p[17]):
	    p4 = p[16].split(';') ##Split multiple references in one line.
	    p5 = p[16].split(';') ##Split multiple databases in one line.
	  elif(',' in p[16])&(',' in p[17]):
	    p4 = p[16].split(',') ##Split multiple references in one line.
	    p5 = p[17].split(',') ##Split multiple databases in one line.
	  else:
	    ListProv.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' "' + 'unclear origin' + '"^^xsd:string')
	    ListErrors.append("CHECK: Multiple Database Provenance And Pubmed IDs detected, check original data for: "+ p[0] + " : " + p[16] + ', ' + p[17] + ' ' + p[18] + '\n')  
	  ##Pubmed IDs:
	  p4 = [x.strip(' ') for x in p4] ##strip whitespaces (if existing)
	  for y in [0,(len(p4)-1)]:
	    if(p4[y].isnumeric()):
	      p4 = ['pubmed:' + s for s in p4] ##add prefix for each item in list
	      ListPMID.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " " + ', '.join(p4))
	  countRefs = countRefs + len(p4)
	  ##Database names:
	  try:
	    ListProvenance = []
	    p5 = [x.strip(' ').lower() for x in p5] ##strip whitespaces (if existing)
	    for z in [0,(len(p5)-1)]:
	      if(p5[z] in ListSupportedDatabases): ##check for original name first
	        p5[z] = '"' + p5[z] + '"^^xsd:string' ##add suffix for each item in list
	        ListProvenance.append(p5[z])
	      elif(p5[z] in ListSupportedDatabasesAlternatives):  ##Check to convert alternative name to official name.
	        for key in Dict_SupportedDatabases:
	          if key.lower() == p5[z].strip().lower():
	            p5[z] = Dict_SupportedDatabases[key] ##Convert alternative name to official to item
	            p5[z] = '"' + p5[z] + '"^^xsd:string'##add suffix for item
	            ListProvenance.append(p5[z])
	      else:
	        ListProvenance = []
	        ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')  
	  except NameError:
	    continue
	  ListDatabase.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' ' + ', '.join(ListProvenance))
	  try:
	    countProv = countProv + len(p5)
	  except NameError:
	    countProv = countProv
	####Second scenario, only pubmed is available:  
	##Option 1: Pubmed contains 1 value; database name contains 0 value
	elif (p[16].isnumeric())&((p[17].strip()=='-')|(p[17].strip()=='NA')|(p[17].strip()=='')):
	  ListProv.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " pubmed:" + p[16].strip( ))
	  countRefs = countRefs + 1
	##Option 2: Pubmed contains more than 1 value, database name contains 0 value
	elif((';' in p[16])|(',' in p[16]))&((p[17].strip()=='-')|(p[17].strip()=='NA')|(p[17].strip()=='')):
	  if(';' in p[16]):
	    k2 = p[16].split(';') ##Split multiple references in one line.
	  elif(',' in p[16]):
	    k2 = p[16].split(',') ##Split multiple references in one line.
	  else:
	    ListErrors.append("CHECK: Multiple separators for Pubmed IDs detected, check original data for: "+ p[0] + " : " + p[16] + ' ' + p[18] + '\n')  
	  k2 = [x.strip(' ') for x in k2] ##strip whitespaces (if existing)
	  ##TODO: build in test to see if all values are numeric!
	  k2 = ['pubmed:' + s for s in k2] ##add prefix for each item in list
	  ListProv.append(p[0].strip( ) + '_measurement' + '\t' + 'dcterms:references' + " " + ', '.join(k2))
	  countRefs = countRefs + 1
	####Third scenario, only Database name is available:  
	##Option 1: Pubmed contains 0 value; database name contains 1 value
	elif(all(x.isalpha() or x.isspace() for x in p[17]))&((p[16].strip()=='-')|(p[16].strip()=='NA')|(p[16].strip()=='')):
	  if(p[17].strip().lower() in ListSupportedDatabases): ##check for latin name first
	    ListDatabase.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' "' + p[17].strip( ).lower() + '"^^xsd:string')
	    countProv = countProv + 1
	  else:
	    ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')
	##Option 2: Pubmed contains 0 value; database name contains more than 1 value
	elif((';' in p[17])|(',' in p[17]))&((p[16].strip()=='-')|(p[16].strip()=='NA')|(p[16].strip()=='')):
	  if(';' in p[17]):
	    l2 = p[17].split(';') ##Split multiple references in one line.
	  elif(',' in p[17]):
	    l2 = p[17].split(',')
	  l2 = [x.strip(' ').lower() for x in l2] ##strip whitespaces (if existing)
	  ListProvenance = []
	  for z in [0,(len(l2)-1)]:
	    if(l2[z] in ListSupportedDatabases): ##check for original name first
	      l2[z] = '"' + l2[z] + '"^^xsd:string' ##add suffix for each item in list
	      ListProvenance.append(l2[z])
	    elif(l2[z] in ListSupportedDatabasesAlternatives):  ##Check to convert alternative name to official name.
	      for key in Dict_SupportedDatabases:
	        if key.lower() == l2[z].strip().lower():
	          l2[z] = Dict_SupportedDatabases[key] ##Convert alternative name to official to item
	          l2[z] = '"' + l2[z] + '"^^xsd:string'##add suffix for item
	          ListProvenance.append(l2[z])
	    else:
	      ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] +  '\n')
	  ListDatabase.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' ' + ', '.join(ListProvenance))
	  countProv = countProv + len(l2)
	####Fourth scenario, no valid entry is available:
	else:
	  ListProv.append(p[0].strip( ) + '_measurement' + '\t' + 'dc:source' + ' "' + 'unclear origin' + '"^^xsd:string')
	  ListErrors.append("CHECK: Name for Database Provenance contains incorrect symbols, check original data for: "+ p[0] + " : " + p[17] + ' ' + p[18] + '\n')  

ListQC.append("Data format for PMIDs Provenance correctly loaded for " + str(countRefs) + " values. \n\n")
ListQC.append("Data format for Database Provenance correctly loaded for " + str(countProv) + " values. \n\n")

AllDict = {}			

##Connect all List data in a Dictionary
##All items are separated with ; (since each line has his own SER_ID.
for itemListSER in ListSER_ID:
	(key, val) = itemListSER.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')

##Finalize statements with type info (if Rhea ID exists):	
for itemListSERtype in ListSER_ID_type:
	(key, val) = itemListSERtype.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' .')	

unique_ListUniprot = list(dict.fromkeys(ListUniprot))

for itemListUniprot in unique_ListUniprot:
	(key, val) = itemListUniprot.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')		

#Remove duplicates
unique_ListRheaID = list(dict.fromkeys(ListRheaID))

for itemListRheaID in unique_ListRheaID:
	(key, val) = itemListRheaID.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')		

##Finalize statements with type info (without Rhea ID, with equation):	
for itemListRheatype in ListRheaID_type:
	(key, val) = itemListRheatype.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' .')		

for itemListSubstrate in ListSubstrate:
	(key, val) = itemListSubstrate.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')	

for itemListKm in ListKm:
	(key, val) = itemListKm.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')		

for itemListKcat in ListKcat:
	(key, val) = itemListKcat.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')	

for itemListKcatKm in ListKcatKm:
	(key, val) = itemListKcatKm.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')		

for itemList_pH in List_pH:
	(key, val) = itemList_pH.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')

for itemListTemperature in ListTemperature:
	(key, val) = itemListTemperature.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')	

for itemListAdditionalConditions in ListAdditionalConditions:
	(key, val) = itemListAdditionalConditions.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')	

for itemListOrganism in ListOrganism:
	(key, val) = itemListOrganism.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')	

for itemListPMID in ListPMID:
	(key, val) = itemListPMID.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val + ' ;')		

##Last item should end with a .
##Scenario 1 and 3: both PMID and Database name are available as provenance; or only database name.
for itemListDatabase in ListDatabase:
  (key, val) = itemListDatabase.strip('\n').split('\t')
  AllDict.setdefault(key, [])
  AllDict[key].append(val + ' .')

##Scenario 2: only PMID is available as provenance (no database)
for itemListDatabase in ListProv:
  (key, val) = itemListDatabase.strip('\n').split('\t')
  AllDict.setdefault(key, [])
  AllDict[key].append(val + ' .')		

##Remove duplicates for linked lists:
unique_ListLinkUniprot = list(dict.fromkeys(ListLinkUniprot))
unique_ListLinkRheaID= list(dict.fromkeys(ListLinkRheaID))
unique_ListSubstrateIDs= list(dict.fromkeys(ListSubstrateIDs))

## Add UniProt proteins for interoperability:  
for itemListLinkedUniprot in unique_ListLinkUniprot:
	(key, val) = itemListLinkedUniprot.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val)

## Add links between UniProt IDs and measurement Enzymes:
for itemListProteinsLinks in ListProteinsLinks:
	(key, val) = itemListProteinsLinks.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val)	

## Add Rhea interaction IDs for interoperability:  
for itemListLinkRheaID in unique_ListLinkRheaID:
	(key, val) = itemListLinkRheaID.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val)	

## Add links between Rhea IDs and measurement Interactions:
for itemListInteractionLinks in ListInteractionLinks:
	(key, val) = itemListInteractionLinks.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val)	

##Add ChEBI IDs for substrates ListSubstrateIDs  
for itemListSubstrateIDs in unique_ListSubstrateIDs:
	(key, val) = itemListSubstrateIDs.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val)	

## Add links between ChEBI IDs and measurement Interactions:
for itemListSubstrateLinks in ListSubstrateLinks:
	(key, val) = itemListSubstrateLinks.strip('\n').split('\t')
	AllDict.setdefault(key, [])
	AllDict[key].append(val)	

#########
  
# # Go to output folder
dir_code = os.getcwd()
parentPath = os.path.dirname(dir_code) #go up one directory
outputFolder = parentPath + "/Output"
os.chdir(outputFolder) ##Update directory to folder to store output data
dir_code = os.getcwd()

# # Empty the data file before writing new content:
open('RDF_Kin_Data_2025-July.ttl', 'w').close()

# # open a file for writing:
RDF_Kin_data = open('RDF_Kin_Data_2025-July.ttl', 'wb')

# #First, print the prefixes from existing databases
##.encode() needed to write to files in Python 3.x (compared to 2.x)
RDF_Kin_data.write("@prefix SER: <http://bigcat-um.github.io/KinRDF/kin#> . \n".encode())
RDF_Kin_data.write("@prefix dc: <http://purl.org/dc/elements/1.1/> . \n".encode()) 
RDF_Kin_data.write("@prefix rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#> . \n".encode()) 
RDF_Kin_data.write("@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> . \n".encode()) 
RDF_Kin_data.write("@prefix wp: <http://vocabularies.wikipathways.org/wp#> . \n".encode()) #From WikiPathways
RDF_Kin_data.write("@prefix rh: <http://rdf.rhea-db.org/> . \n".encode()) #From Rhea
#RDF_Kin_data.write("@prefix RHEA:   <https://www.rhea-db.org/reaction?id=> . \n".encode()) #For website link, not for IRI!
RDF_Kin_data.write("@prefix RHEA:   <https://identifiers.org/rhea/> . \n".encode()) #To link to WPRDF
RDF_Kin_data.write("@prefix CHEBI:   <http://purl.obolibrary.org/obo/CHEBI_> . \n".encode()) #To link to Rhea RDF (which includes ChEBI)
RDF_Kin_data.write("@prefix dcterms: <http://purl.org/dc/terms/> . \n".encode()) 
RDF_Kin_data.write("@prefix xsd:   <http://www.w3.org/2001/XMLSchema#> . \n".encode())
RDF_Kin_data.write("@prefix uniprot:   <https://identifiers.org/uniprot/> . \n".encode())
RDF_Kin_data.write("@prefix uniprotkb:   <http://purl.uniprot.org/uniprot/> . \n".encode())
RDF_Kin_data.write("@prefix up:   <http://purl.uniprot.org/core/> . \n".encode())
RDF_Kin_data.write("@prefix ECcode:   <https://identifiers.org/ec-code/> . \n".encode())
RDF_Kin_data.write("@prefix En_id:   <http://identifiers.org/ensembl/> . \n".encode())
RDF_Kin_data.write("@prefix pubmed:  <http://www.ncbi.nlm.nih.gov/pubmed/> . \n".encode()) #For WPRDF interoperability
RDF_Kin_data.write("@prefix NCBI:  <http://purl.obolibrary.org/obo/NCBITaxon_> . \n".encode()) #For NCBI Taxon IDs
RDF_Kin_data.write("@prefix wd: <http://www.wikidata.org/entity/> . \n".encode()) #From WikiData
RDF_Kin_data.write("@prefix wdt: <http://www.wikidata.org/prop/direct/> . \n".encode()) #From WikiData
RDF_Kin_data.write("@prefix sio: <http://semanticscience.org/resource/> . \n".encode()) #Semanticscience Integrated Ontology
RDF_Kin_data.write("@prefix bioregistry: <https://bioregistry.io/oboinowl:> . \n\n".encode()) #Bioregistry hasDbXref

# #Second, print the information on units (for measurements) SIO_000221 
## Temp wdt:P2076
RDF_Kin_data.write("wdt:P2076 wd:Q47574 wd:Q25267 . \n".encode()) #Temperature unitOfMeasurement degree Celcius (Wikidata Q25267) #unit of measurement (Q47574)
RDF_Kin_data.write("#Temperature unitOfMeasurement degree Celcius (Wikidata Q25267) \n".encode()) #Print comment for clarity
## Km SER:hasKm ; Q61751178
RDF_Kin_data.write("SER:hasKm wd:Q47574 wd:Q105687351 . \n".encode()) #Michaelis constant unitOfMeasurement  millimolar mM (Wikidata Q105687351)
RDF_Kin_data.write("#Michaelis constant unitOfMeasurement degree millimolar mM (Wikidata Q105687351) \n".encode()) #Print comment for clarity
RDF_Kin_data.write("SER:hasKm wd:Q47574 wd:Q61751178 . \n".encode()) #Michaelis constant a Michaelis constant (Wikidata Q61751178)
RDF_Kin_data.write("#Michaelis constant a Michaelis constant (Wikidata Q61751178) \n".encode()) #Print comment for clarity
## KCat SER:hasKcat
RDF_Kin_data.write("SER:hasKcat wd:Q47574 wd:Q6137407 . \n".encode()) #Turnover Number unitOfMeasurement  reciprocal second s-1 (Wikidata Q6137407)
RDF_Kin_data.write("#Turnover Number unitOfMeasurement degree reciprocal second s-1 (Wikidata Q6137407) \n".encode()) #Print comment for clarity
RDF_Kin_data.write("SER:hasKcat wd:Q47574 wd:Q899698 . \n".encode()) #Turnover Number a Turnover Number (Wikidata Q899698)
RDF_Kin_data.write("#Turnover Number a Turnover Number (Wikidata Q899698) \n".encode()) #Print comment for clarity
## Kcat/Km SER:hasKcatKm 
RDF_Kin_data.write("SER:hasKcatKm wd:Q47574 wd:Q116486875 . \n".encode()) #Specificity constant unitOfMeasurement M−1s-1 (Wikidata Q116486875)
RDF_Kin_data.write("#Specificity constant unitOfMeasurement degree M−1s-1 (Wikidata Q116486875) \n".encode()) #Print comment for clarity
RDF_Kin_data.write("SER:hasKcatKm a wd:Q7575016 . \n\n".encode()) #Specificity constant a Specificity constant (Wikidata Q7575016)
RDF_Kin_data.write("#Specificity constant a Specificity constant (Wikidata Q7575016) \n".encode()) #Print comment for clarity

## Ph has no unit.

#Third, print the dictionaries.
##ERPX_number			
for KeySERX, ValueSERX in AllDict.items():
	RDF_Kin_data.write(KeySERX.encode('utf-8') + "\n".encode())
	for item in ValueSERX:
		RDF_Kin_data.write("\t".encode() + item.encode('utf-8') + "\n".encode())
	RDF_Kin_data.write("\n".encode())	
# for i in range(LengthList):
     # RDF_Kin_data.write("This is line %d\r\n" % (i+1))

#Close this file as well
RDF_Kin_data.close()

# # Empty the QC file before writing new content:
open('QC_RDF_Kin_Data_2025-July.ttl', 'w').close()

# # open a file for writing:
RDF_Kin_data_QC = open('QC_RDF_Kin_Data_2025-July.ttl', 'wb')

##Write QC info to file:
if len(ListQC) >0:
  for item in ListQC:
    RDF_Kin_data_QC.write(item.encode())

RDF_Kin_data_QC.write("\nReported Errors in Data: \n\n".encode())

##Write Error info to file:
if len(ListErrors) >0:
  for item in ListErrors:
    RDF_Kin_data_QC.write(item.encode())
    
RDF_Kin_data_QC.write("\nPotential Curation Required: \n\n".encode())    
    
##Write Curation info to file:
if len(ListCuration) >0:
  for item in ListCuration:
    RDF_Kin_data_QC.write(item.encode())

#Close this file as well
RDF_Kin_data_QC.close()

# # Go back to parent folder before finishing the script:
dir_code = os.getcwd()
parentPath = os.path.dirname(dir_code) #go up one directory
dir_code = parentPath

#Remove all variables in working directory:
for element in dir():
    if element[0:2] != "__":
        del globals()[element]
del element
